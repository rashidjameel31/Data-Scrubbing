import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from pathlib import Path
import PySimpleGUI as sg
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
#from selenium.webdriver.chrome.service import Service as ChromeService
import time 
from sys import exit
#from webdriver_manager.chrome import ChromeDriverManager
class bcd:
    def __init__(var,label,down):
        cntr = True
        banerr=[]
        err_wb = openpyxl.Workbook()
        err_ws = err_wb.active
        var.label = label
        var.down = down
        print(down)
        WebDriverWait(inst, 20).until(EC.element_to_be_clickable((By.ID, 'Actionsspan')))
        inst.find_element(By.ID,'Actionsspan').click()
        inst.find_element(By.LINK_TEXT, "Banner View").click()

        inst.implicitly_wait(2)
        inst.switch_to.window(inst.window_handles[2])
        inst.implicitly_wait(2)
        i = 2
        while(cntr):
            try:
                WebDriverWait(inst, 10).until(EC.element_to_be_clickable((By.ID, 'downloadLink')))
                #inst.find_element(By.XPATH, "//table[@id=\'ATTACHMENTS_FILELIST\']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[4]/a").click()
                inst.find_element(By.XPATH, "/html/body/div/table/tbody/tr["+str(i)+"]/td[2]/span/a").click()
                #inst.find_element(By.ID, "downloadLink").click()
            except:
                try:
                    inst.find_element(By.XPATH, '//*[@id="rev"]/option[2]')
                    inst.find_element(By.XPATH, "/html/body/div/table/tbody/tr["+str(i)+"]/td[2]/span/a").click()
                except:
                    if(i == 2):
                        banerr.append(var.label)
                        banerr.append("Download link not found")
                        err_ws.append(banerr)
                        err_wb.save('Error.xlsx')
                    cntr = False
            time.sleep(2)
            i = i+1
            while (True):
                chrome_temp_file = sorted(Path(var.down).glob("*.crdownload"))
                #downloaded_files = sorted(Path(var.down).glob('*.*'))
                print(chrome_temp_file)
                #print(downloaded_files)
                if len(chrome_temp_file) == 0:
                #if (len(downloaded_files) >= 1):
                    break
        inst.implicitly_wait(2)
        inst.close()
        inst.switch_to.window(inst.window_handles[1])
        
class scrap:
    def scr(self):
        '''wb = openpyxl.Workbook()    
        ws = wb.active'''
        f=[]
        try:
            f.append(inst.find_element(By.ID, "col_1001").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1081").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1084").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1002").text)
        except:
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1082").text)
        except:
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1068").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1004").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1017").text)
        except :
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1016").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_12089").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008074").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2023").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2029").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2092").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2090").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2091").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2007").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2024").text)
        except :
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2021").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008063").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1080").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1420").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1271").text)
        except :
            f.append("NA")
            pass
        return list(f)

class nfd():
    def __init__(var,label,down):
        flag = True
        err_wb = openpyxl.Workbook()
        err_ws = wb.active
        var.label = label
        var.down = down
        inst.find_element(By.LINK_TEXT, "Attachments").click()
        try:
             WebDriverWait(inst, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@id=\'ATTACHMENTS_FILELIST\']/tbody/tr[2]/td[2]/div/div/table/tbody/tr[2]/td[4]/a")))
        except:
            lir = []
            lir.append(var.label)
            lir.append("No Attachment Found")
            err_ws.append(lir)
            err_wb.save('Error.xlsx')
            flag=False
        if flag == True:
            try:
                for i in range(2,8):
                    time.sleep(2)
                    inst.find_element(By.XPATH, "//table[@id=\'ATTACHMENTS_FILELIST\']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[4]/a").click()
                    time.sleep(3)
                 #inst.find_element(By.XPATH, "//table[@id=\'ATTACHMENTS_FILELIST\']/tbody/tr[2]/td[2]/div/div/table/tbody/tr[3]/td[4]/a").click()
                    #inst.find_element(By.XPATH, "//table[@id=\'ATTACHMENTS_FILELIST\']/tbody/tr[2]/td[2]/div/div/table/tbody/tr[4]/td[4]/a").click()
                    #inst.find_element(By.XPATH, "//table[@id=\'ATTACHMENTS_FILELIST\']/tbody/tr[2]/td[2]/div/div/table/tbody/tr[5]/td[4]/a").click()
                    while (True):
                        chrome_temp_file = sorted(Path(var.down).glob("*.crdownload"))
                        #downloaded_files = sorted(Path(var.down).glob('*.*'))
                        print(chrome_temp_file)
                        #print(downloaded_files)
                        if len(chrome_temp_file) == 0:
                        #if (len(downloaded_files) >= 1):
                            break
            except:
                pass
            inst.implicitly_wait(2)
            
class co():
    def scr(self):
        f=[]
        try:
            f.append(inst.find_element(By.ID, "col_1047").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1030").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1069").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1060").text)
        except:
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1052").text)
        except:
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1053").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1049").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1003").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_3742").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "R1_1099_0").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "R1_1050_0").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1061").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1051").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_3743").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008074").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2092").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2020").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2023").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2018").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2091").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2090").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2017").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2021").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2019").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1271").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2024").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1331").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2007").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2008").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1080").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008063").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008073").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008064").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008065").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008066").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008067").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008068").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008069").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008070").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008071").text)
        except:
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008072").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1567").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1568").text)
        except :
            f.append("NA")
            pass
        return list(f)

class cse():
    def scr(self):
        '''wb = openpyxl.Workbook()    
        ws = wb.active'''
        f=[]
        
        try:
            f.append(inst.find_element(By.ID, "col_1047").text)
            f.append(inst.find_element(By.ID, "col_1030").text)
            #f.append(inst.find_element(By.ID, "col_1052").text)
            f.append(inst.find_element(By.ID, "col_1061").text)
            f.append(inst.find_element(By.ID, "col_1051").text)
            f.append(inst.find_element(By.ID, "col_3743").text)
        except :
            f.append("NA")
        inst.find_element(By.LINK_TEXT, "Workflow").click()
        try:
            for i in range(2,20):
                #inst.find_element(By.XPATH, "//table[@id=\'ATTACHMENTS_FILELIST\']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[4]/a").click()
                
                status_path = str(inst.find_element(By.XPATH, "//table[@id='CHANGETABLE_SIGNOFF_HISTORY']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[4]").text)
                print(status_path)
                if status_path == "Approval":
                    f.append(str(inst.find_element(By.XPATH, "//table[@id='CHANGETABLE_SIGNOFF_HISTORY']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[10]").text))
                
                elif status_path == "Released":
                    f.append(str(inst.find_element(By.XPATH, "//table[@id='CHANGETABLE_SIGNOFF_HISTORY']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[10]").text))
                elif status_path == "Implemented":
                    f.append(str(inst.find_element(By.XPATH, "//table[@id='CHANGETABLE_SIGNOFF_HISTORY']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[10]").text))
        except:
            pass
        return list(f)

class bcdval():
    def __init__(var,label,down):
        f=[]
        try:
            f.append(inst.find_element(By.ID, "col_1047").text)
            f.append(inst.find_element(By.ID, "col_1030").text)
            #f.append(inst.find_element(By.ID, "col_1052").text)
            f.append(inst.find_element(By.ID, "col_1061").text)
            f.append(inst.find_element(By.ID, "col_1051").text)
            f.append(inst.find_element(By.ID, "col_3743").text)
        except :
            f.append("NA")
        inst.find_element(By.LINK_TEXT, "Workflow").click()
        try:
            for i in range(2,20):
                #inst.find_element(By.XPATH, "//table[@id=\'ATTACHMENTS_FILELIST\']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[4]/a").click()
                
                status_path = str(inst.find_element(By.XPATH, "//table[@id='CHANGETABLE_SIGNOFF_HISTORY']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[4]").text)
                if status_path == "Approval":
                    f.append(str(inst.find_element(By.XPATH, "//table[@id='CHANGETABLE_SIGNOFF_HISTORY']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[10]").text))
                
                elif status_path == "Released":
                    f.append(str(inst.find_element(By.XPATH, "//table[@id='CHANGETABLE_SIGNOFF_HISTORY']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[10]").text))
                elif status_path == "Implemented":
                    f.append(str(inst.find_element(By.XPATH, "//table[@id='CHANGETABLE_SIGNOFF_HISTORY']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(i)+"]/td[10]").text))
        except:
            pass
            
        cntr = True
        banerr=[]
        err_wb = openpyxl.Workbook()
        err_ws = wb.active
        var.label = label
        var.down = down
        print(down)
        WebDriverWait(inst, 20).until(EC.element_to_be_clickable((By.ID, 'Actionsspan')))
        inst.find_element(By.ID,'Actionsspan').click()
        inst.find_element(By.LINK_TEXT, "Banner View").click()

        inst.implicitly_wait(2)
        inst.switch_to.window(inst.window_handles[2])
        inst.implicitly_wait(2)
        i = 2
        while(cntr):
            try:
                WebDriverWait(inst, 10).until(EC.element_to_be_clickable((By.ID, 'downloadLink')))
                inst.find_element(By.XPATH, "/html/body/div/table/tbody/tr["+str(i)+"]/td[2]/span/a").click()
            except:
                try:
                    inst.find_element(By.XPATH, '//*[@id="rev"]/option[2]')
                    inst.find_element(By.XPATH, "/html/body/div/table/tbody/tr["+str(i)+"]/td[2]/span/a").text
                except:
                    if(i == 2):
                        banerr.append(var.label)
                        banerr.append("Download link not found")
                        err_ws.append(banerr)
                        err_wb.save('Error.xlsx')
                    cntr = False
            time.sleep(2)
            i = i+1
        inst.implicitly_wait(2)
        inst.close()
        inst.switch_to.window(inst.window_handles[1])
        return list(f)



logo = b'iVBORw0KGgoAAAANSUhEUgAAA3wAAAM6CAMAAADty7FhAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAEjUExURQAAACBggDBQjytVijBQjy1TjDBQjy5SjTBUjy5RjjBTjy5RjjBSjy9RjjBSjy9RjjBSjy9TkC5Sjy9TkC9Sjy9SkC9Sjy9SkC9Tjy9SkC9Tjy9SkC9Sjy9SkC9Sjy9SkC9SjzBPhDBRijBRjTFOfzFOgjFPhzJMeTJMfDNIbjNJcTNKdDNLdzRHaTRIbDVEYTVFZDVGZjZCWzZDXjc/UDdAUzdBVjdCWTg+Szg/Tjk7QDk8RTk8SDo5Ojo5PTo7Qzs4OEE+PkdERE1LS1RRUVpXV2BdXWZkZGxqanJwcHh2dn99fYWDg4uJiZGPj5eWlp2cnKOhoamoqK+urrW0tLu6usLBwcjHx87NzdTT09ra2uDg4Obm5u3s7PPz8/n5+f///5t2oV0AAAAgdFJOUwAIEBggKDA4QEhQWGBocHiAh4+Xn6evt7/Hz9ff5+/3v+H4wgAAAAlwSFlzAAAywAAAMsABKGRa2wAAQuhJREFUeF7tnWlDFLkWhlll3xcRBQpQdqFZBEGWHr3OoOOMzKijM278/19xO8mp7gZ6qepKqs5Jvc+HO962uypVyetJcpa0Af509pboHy0zPVNmkL4CALCDktvI6OhESV4LQUPud9JPAAAt06msW3O53WIexg+AFmlXRm56Zo7EFB8YPwDi0dM7NDo5M0sKSsJ8P10SANCYnoHR+w9JOFFYXV3dLpTZPy5zuELfmOygSwMAahNVdmurq+slnZX0dVFsxPk6/WCuj+4AALhF7+DoTBPZraxuFnaOj1+QsKJxtES/hvED4DZdA6Mzj0ghtVheXd0tPDs+ITXF5WSLrgPjB0CF9t7R6bq7mIulhdze8SlpKAFl4zfWTvcFINfcG5yot5O5uFo4Oibl2KBs/B720M0ByCmdfaN1POW2ZReyt0w3gPED+aVneKrmAs+V7IgXa3QfGD+QS3rHHpAEbvB426XsQnZC4zdCrQEgJ/SM3KfBX83K5k4KujOUjd+DbmoTAN5zb3j67hJveXX3WavugxbZoVsvwPiBPNA1NHXXmfB4e8+CCyE+z59QA2D8gOd0Dkze2VxZ2Swcn5EW0udsl5qxMExtBMA72vvH70SLLa4fpTzRvEvZ+CHRCHhJx+CdRd7Sxl68oExXlI0fsmyBd3QNzdDwDlne2s9kiVeH4zDRCMYP+ET3yO2YsbVCas6EqJQTjWD8gC/0jN/aX3n89CC7zZVGHCDLFnhE/8RNl8LS+uE5DXWGIMsWeELHwNQ8DWbD4ja7ueZtDpBlC8TTPjBNw5h4UuC0vVIXZNkC4fRO3rR5q/uNq6pwopxlO4FEIyCN7rEbOyzLm9k70WOBLFsgk86hG16FpfVDnjubDUGJCSCPgSkatRoBGyx1uECWLRBFz0T1Qm9p+zmNZJGUs2xh/AB3ukZvLPQ2DmkQiwVZtkAE7UM3KkGs7TF2pEcnNH7IsgVs6b4x3Vzc5ZGnYIEXyLIFrBmorsKytC51i6UmlSxbGD/Ajc6R6pXelkS3QmPKWbYzSDQCnOiZrEqOfbIvy5UeEWTZAn60D1Ztsixte7PQuwNKTABedI1VbbI88WN3sx7IsgWM6KvOWNg4oEHqL+Us2/tINAJZ0jFUtcmysuvlSu82yLIFDOisnm+uHdHg9B+UmAAZ0zVe2d9cXvd3k6UG55v03DB+IAO6Jmn8lVjc93qTpRY4yB1kxb2qdKEt/zdZaoAsW5AJvVUbnNtyakJYBlm2IHX6KuWml5/mYn+zDier9Bpg/EAq9FcqQywV8iw9BbJsQXoMVtx6Kzu522W5S+UgdyQaAZe0D1ekt7jvXdJCayDLFqTAUKXY+5P8ONSbghITwDUDFav3+BmNO6BBli1wSV9lm2XVq/x0K+Agd+CMnopzYVN0DUBXIMsWuOFexaW+LuJ8kyxAli2wT1UM5yakV5+zp/SWYPyAHTormQtY6zWhkmUL4wcS0zFalt7jXEZPx6NSYgKJRiAZ7cPlVNlF8QXf0wFZtsAKFZ/6yh6NLdAMlJgAyel9SKMoWNpBIFkMDpFlCxLRWc6VXd5F+HQ8Kge599LbBCAy7SPlfZZc5+u1CrJsQav0lWecG7nNUk8GSkyAlugqx7OswafeMnvIsgVxaS979lbgXUhCJcsWxg9Eor+cN4R9lqQgyxbEoLt8tOVqrkrgOgJZtiAqHWPlGSeSZe0A4wciMRAGtCzvwqluC2TZguZ0l9Nlt+BesEjlIPchetMA3KTsVV9E8oJlkGULGtETnum8XMCM0zooMQHq0j5GgyPYRCyZE54jyxbUpCcMJnuCGacrcJA7qEHZ7C3v0EABLkCWLbhN2eytwavuFmTZghvA7KUJjB+oALOXLpUsWxi/nAOzlz7lLNtxJBrlmb4wfwFmL0WQZQvaOsIy1Ev7NCxAOqDERN7pCYOot+BWT5sLZNnmmlHq/iUccZkFlYPcqT9AbugMIzlh9jICWbZ5pZ9qwMPsZQiybPNI+4Tp9WAVZi9LXjymfoDxyw3doV8dvr2MqWTZwvjlgyHKmV3E2c7ZgyzbPNERlsPdQFlADiDLNj+Ezr1l7LRwAcYvJ4ShnI8RTsaHM2TZ5oCu0Ln3FFVaWFFONJpGopGnDITOPZSK4AaybP0Gzj3WIMvWY8rxZHDu8eR8kzoIxs83wl1OOPf4Uk40gvHzitCxDuceZ5Bl6yHhcm95j3oZMAVZtr4RLvdWMOVkD4yfX4TLvTXsckqgkmUL4yeecLm3Dce6DMpZtrNINJINlnsCQZatF2C5JxKUmPAALPekUjAdB+MnFiz35FI5yL2LehMIAss90SDLVjAdWO4JB1m2UumiKklY7skFxk8m92irBcs90ZQTjWD85NBntlqw3JMOsmzFMWg6bAnLPfkgy1YWVCZp8YL6D0gGxk8Q7XTs3mPk7nnCIbJshdBx33TUFrZavKFykDsSjTjTNWu6aZv6DXgBsmwFcI9OWUeZJM9Ali17ek1pThSD95AjZNmyZtC491AV10vKiUYwfgwZMX2zeEq9BTwDWbZsoSyGx3DveQuybHnSTi6GVbj3fAbGjyGh9jbg3vOb03KWLYwfE0Lt7VIXAW+pHOSORCMWhNpDFkMeQJYtJ8Ksdbj38gGybPnQYULK4FrPD88XdZfD+GVNqD241nNEOdFofoCGAcgAaC+fIMs2e6hSErSXO5BlmzVdJo0B4Zx5BMYvU0LtIZwzl1SybGH8UgfayzvIss0Kqs4J7eUYZNlmwz2TOrsC7eUaGL8MIO2hQmDeOUGWbdp0mTkntAcqB7mP0uAATqG9FmgPlECWbZp0GN86tAcMyLJNDYopW4H2AAHjlxKUvwcfA6hQybKF8XMIaW8Z2gPVIMs2BcxRKIilBreoZNkO9Zag4QIsQjUCoT1wh7LxC5mZmRkd7euioQMSMmreKvLWQQ3OwkSjm8zPjA3eowEEWobqUqNWEqhNOdHoLg8mhzEXTQAd+VygNw3AbcpZtjWZnxxADGhr9Js3iPqcoAEHq4Y6NnBhCvprgT5zDtE6vWQAmnB6fLyzvapHzQ2mB5H9Ho8eo70NerEAROTFYeGOHZxG1c8YdJskIhy3Dlri/Hhn44YCH0F+Uek0iQyr0B5onePtav1BftFoN0XhH+MMMJCM46dhwesSkF8UpvS7gvaABU6r9Dc3jL3PJozpF4UkImCJ48oe6NwwDTJQkwH9lpDIAOxxuqkHlQJJgA3oNU6GZ/TaALBBRX5IAqxLl3Ey7NA7A8ASF+VYNBi/2lDFFgS2APuU5QfjV4v2Gf1y4OADTignAcL43cVkri/CyQDcgPIvdTEZfEtwMgBnlI3fJHx+1ZgsouVjek0AOKBs/GaQ7VCBMhmQuQ7cEhq/hyj4EkJHMiB7Frjm7KkeasE8ir0YqDT1Jr0fABxypAdbsNBPoy/nTOu38RhOBpAGB3TWH2I9SwzrV4HjUEBKnFKywwQNwBxjNlsQTQ1S4+SxFl8wSUMwt3SazRaUxwXpcUaHvOfd9pmosqf0VgBIhW097HJ+yq2pC7+GzRaQLvt64OV616VXvwFElYHUIfXlt7oLpfDhLCKQPgU99oI+Got5g2qV4UwGkAUmx28hp8eqjOunX6V3AUC6bOjxl89IM5PKsHJCrwKAdCGPw3wOo6y7jXf9Ob0JANLmzHjbH+Quv6/d1GzZp/cAQPqcG/WN05jMDaZuBFIZQJZcmDjPnKU4mNNnUbMFZMuLZTUO87Xso3DqF/QKAMgIk9+Xp2VfuzkJDOHUIHOMuy9Hyz6z4EOBXJA9Z6auS27izPr04yKcGnDALPsWclJM1+TwIZwa8MAs+x7mY9lnirZgwQeYYJZ9uUjuM14GePgAF8yybyEHQZ5d2suAkE7AB7Pse0Aj1GNMHhFy+AAjdvSgHKIh6i2mcMQ2PTQALNBBngudNEg95Z6edD6BlwGw4lQNy2CaRqmfUC4D8ogAM8whDgM0Tr3EJK+jcATgxtmKGplzHh8eZqqVIbQF8OOZHpv+1tHt0KEtyGUAHDElXXporHrHlH48nIEJOHKijy+apbHqGya0ZYueFQBe7Onx6WdWuymRu4TQFsAU7ezz0/SZeOpDelAAuGH2XHw8v8GU6UQCLeDLqhqic/7lFrXrnc4VVEwCfDFxLv6ZPuNex6QTcGZTDdJ530I872ntYacTsOZU5xb5Vk1pVj3UMipHAN7oEE/PshuG1TMFO/SEADDlRJu+SRq2XtCpE4keI6YTcMdkN/hUUcLElSGRCLDHZDd4FF9t6nQiex0IQB/W7s+qzxSHh4sPSMCYPm82PMfU08DFB2RQUKPVF9NnyrbAxQdkYDY8R2j0CkfXCkQGLZCC3vD0I8JzSD0KyrYAMRjT50OEpzkVBbUCgRz02Q0+mD5zFN8xPRYA/DHJDfLLCPbo50AWH5CETm6Qn9KuA6pxKgoQhTF90qu5mPR11CsDstAp7fdpEAvFVIdHQDUQhqnmIvukaJNJ9IyeCAAp6OMyRceYmbota/Q8AIhBh1fPS/Y2mLotp/Q8AIjhXDvaBZ+WaVJocfg6EMi2GruCvQ3av55JUOfbK778eVnmNTXXb+i5GfIntbAWxtsg9tgUU7DsKT1Mqny8FsO/X758Ko2DN5f/o7Z7Bz0oQ75RC2uypkbvFI1lcejy8MuZ+NcFia+Kb18+Xr3zT4T0dAxpKL5DNXylpvUZ//o+PUq6yBRfyNdPV5cv6Uk8gJ6KIQ3FZzLahab1mcCybPzrssWn+f756q0fRpAeiCENxVfcVQN4jkazLMxhfEf0ICnjgfg03z68/YUeSS70LAxpLL4TPYL7aDxLwvjXH9NzpI0v4ivx858/hBtAehCGNBafyW2QuOWSbWCZR+JTfP1Lsv7oIRjSRHxmy6WDRrQcOvUptJn51z0TX4nPv9OjyYOegCFNxHemD2mXF+WScWCZf+IrjZT3r+jphEHtZ0gT8Zkolwc0pMVgAsuyy1/3UXwlPv5GzycKajxDmonvuRrF4hKLtOHLxr+u8VR819dffqUnFAQ1nSHNxGcSi8ZoUAvBGL4MqwV6K77r6w/iJp/UcIY0Fd+OGsfCXH3a8GXkX9d4LL7rH++Fuf6o3QxpKj6Brj5j+LIJLDP4LL7SmHlLjykDajVDmoqvuKVGsihXX9YrPt/FV1r6SUpGojYzpLn4xLn6jOHLJJUoxHfxXf94R08qAGoyQ5qLT5yrL3vD57/4rq8/isl7oAYzpLn4jKtvhkY2f0xUZ6aGLw/iu/4qxetA7WVIBPGZhHYxWX06qjNbw5cL8V3/bFQFgRHUXIZEEJ9x9UnJ6uNg+PIhvuvrzyKmntRYhkQRn3b1SQkx42D48iK+6/8keNyprQyJIr4LNZ6DLhrdvGFh+HIjvutvAnKNqKkMiSI+U0hJxryTheHLj/iuv/HfdqGWMiSS+HTt6oc0vFnDw/DlSHzXP9irjxrKkEjiMyFmElIbTAJ7FnVyb5Aj8V3/eEMPzRVqJ0Miic8cFzZKA5wxxvBlXyA+T+K7/slcfdRMhkQT354a049ohDPGGL7sT0bJlfi4q49ayZBo4jNnpvAvHP9INZPBySj5Et/1D9Zx1tRIhkQTn0ltYJ9Sa4pUMzgSLGfiu/7G2dtObWRIRPHp1Ab28059OgOHI8HyJr7rL4wTbKmJDIkovjMJ885u1cbggJqcJbkT3/UHenKGUAsZElF8IuadOpcoqyLVN8if+K75RllTAxkSVXxHamDznne2Z149okwOxXfNdsuT2seQqOITsN9pIssyLJtUIY/iY7vlSe1jSFTxCZh3PlQtzDyyTJNH8V3/Sw/PDWoeQyKLj72f3fgZMo8s0+RSfNfv6emZQa1jSGTxmfhOxvNONn6GEvkU30+eE09qHUMii8/Ed/Kdd3ap5mV2Jtgt8im+6/9YevuocQyJLj497+SbV6T9DE+orVmTU/FdX9Hzs4LaxpDo4uOdV9SuD+Rj4WcokVfx/eSY3EdtY0h08fHOK2LkZyiRV/Fdf2U48aSmMSSG+FjPO2dV43j4GUrkVnzXf9EbYAS1jCExxMd53tmn28Ygn8GQX/H95FfPjFrGkBji4zzv1H6GLWpn9uRXfNef6BXwgRrGkDji0/U7Wc47WfkZSuRYfNfs9lyoXQyJIz5Tv5PjvFP7GRapmQzIs/j+oXfABmoXQ+KIj239Tl5+hhJ5Ft/1Jb0ELlCzGBJLfFzrxuuwzuVzaiUDci2+/+glcIGaxZBY4uNaN15vt2xQIzmQa/FdMzs1mlrFkFjiYzrv7NBZtBzKR4TkW3zMQjypVQyJJz6e884h1aoVaiIL8i0+ZiUlqFEMiSc+nvPOB6pRu9REFry+DLlS/PXlK73ufMBr1Uc9oflD98fV5y/fqalp8vXLX/rub6kxl79RCyPC8ZxM4+RjE91Sh1/fffpBneA/MUdVJvzv97//pea65/vH35P7PwtqnDObd46qNrEoWtaMy8/UF77DL8ylNi/fp2IBP9kJPHihBjqz89l1jXhGTr5GXOZj/skwwrMOv/xFTXbHZ2sxP3reOUTDngU9qkWZH4cZlZf/UZf4Dd8qnne4oia7wmLEj553ztC4Z8GEahGfmOpmvPpGneI1X+lpJfCB2uyGzxb9LuzmnSa07JCaJ4C31Ct+wy3GrAGvflKbXWB3As5t3jmg2rPEJYU9Cq4mnp9pA/sOf15dXf3z5UuqJpddeHUD/qY2u+Aj3cMO3OadOrRsmxongj+oX2zTtJ9fXl59Tsu/JWfLpVj8ldrsArsJVs/VYOcz7+zUoWXH1DgR/EL9Ypto/8i+enuVyo4r0wq6NXE377QdbrCoRvsgjf3M0aFljDL5ouBo3hl9hvPbR5fLHMMXupcEvlCb7WP75LRdNdzv09jPHB1aVqCmCcFR5Gec5cWr965XgD8ZH5d5G3eLvj/oDrbQ886FDhr8GWPOw7ygpgnhT+oZy8Rc279xHG7DLLGoEe+oyfaxXlOD07xTh5atUcOkcEk9Y5mY4iuNOaeTz7/pLgJwJz7r5p/TvFOHlu1Rw6TgaHcttviKr11GFnNLaG+Ao38NS9AN7MFo3nlPNYVT/YhI/I96xjLxxVcsvndo/P5H9+DPb9Ri63ynG1hkRQ15FvNOPeuUE1pGvKSusUwr4iv+6s7v8I5uwR9H/xrGzZaNxLYa8tM0/jNF73VKm3W6yq1uSXwOY01ba08WOBOfg8N6j9WQX2gnAWSISaMVttdZgrrGMi0O9teucnx/0A3440x8Lpydet45QArIEO1hl7bXWYK6xjKtWpo3rtZ91vfZXSFKfHreOUUKyJD7qh3CPOwK6hrLtDzNc7XTLiapT5T4mMw7TclA7sVbakBdY5nW11iOsrnFePpEia+4pEZ95vNOnU0kLK5TQ11jmQQbHG6CXT7T1dkjS3zrathnPu+cUq1gcyBmDKhrLJNAfG7c/mLS2WWJ70AN+6znnSaHXVQ2EUFdY5kE4iv+Q9ewi5TYalniO9Pzzj5SQUb0qjaIymEPoa6xTBLxuTF9r+nq3JElPjPvnCQVZISunLRODRIFdY1lkojPjemTktggTHx63jmf7bxTB1ULqpxUgbrGMonE58T0SfE1CBOfmXf2kgwyQWZQtYa6xjKJxOfE9NnO43aFMPGZeecE6SATZAZVa6hrLJNMfC4qGjoae9aRJr7s552zqgXygqoV1DWWSSY+F6kWDjJqnCBNfGbe2U9CyACpQdUK6hrLJBOfk7JOQnwN0sSXeXznsLq/wKBqBXWNZRKKz0UVISH5tOLEl3V854y6v8CgagV1jWUSis/Fog/ic0S2+eztUoOqFdQ1lkkoPheLPgmHZJaQJ75s6yjp8BaJQdUK6hrLJBSfi0WfkONS5Ikv27rx2tEgMrylBHWNZZKKz8GiD+JzRabnFekl3xG1RBrUNZZJKj4HtXyF1FASKL4sz2c3Sz6Zjgau4nOQ0A7xucKck9lFckgX0Us+puJ7Q9exiO2jChwhUHzFNaWAEZJDuohe8jEVn4PSsVd0aeZIFN+OUsAjkkO6iF7yMRWfgyEI8TnjRCkgk9QG2Us+puJz4OgTUkJJoviKW0oCWaTUyl7yMRWfg2YlblI6iBTfodJAFiFmspd8XMVnv3A8xOcOk9qQQYiZ7CVffsT3F12ZOSLFZ1IbZkgR6SF8yZcf8WHDxSEmxCx1V5/wJV9+xPeerswcmeIzIWZjpInUEL7ky4/4EOHiEu3qmyNNpIbwJR9X8dk/Kwzic4lx9aVcPVf6ko+r+Og6FhFSuFOo+IyrL+VqEtKXfPkRH1KKnJKFq0/6ko+p+BwMQYjPKcbVl25Wn/QlH1PxXdJ1LCLksAap4is+VUqYJVmkgvglH1PxOcjnQwElt5wqJQQ9JIw00HXil+j2IqGusUxS8V3RdSzyii7NHLHiK64qLaS55TKobrhKdxcJdY1lkorvE13HHj/pytyRK75nSgsLKRZSGlc3lHggbRnqGsskFZ/96mVSjqaVK74zXcAzxYR28fstPMXnIJ3vH7o0d+SKzxRSSjHKRZ8GLbRcroG6xjIJxedgs1NILq1k8Z0sKzWkdmZKt7rbMt1bJtQ1lkkovvd0GYsIqZ8kWXzFTSWH1GpX96u7PaZby4S6xjIJxfeZLmORN3Rp7kgWnz4zJegmcbhmTN1sm24tE+oayyQT38ufdBmLCPGxixZf8bHSwziJwzXT6mYyD8UMoa6xTDLxOXCxX9Ol2SNafPtKD2l5G+bUzZ7TnWVCXWOZZOJzMOuU4mmQLb4zveWSjunrULcKzujOMqGusUwi8b1yMOv8TNdmj2jxmQDPhQ7Sh1P61K2e0H2FQl1jmUTic3Eu7Qe6Nntki894G4ZJH04ZUXfaoPsKhbrGMknE58LwScljly4+Y/rm0kjrm1J32qfbCoW6xjJJxOfC8EnJaRAvPnNiURqm76G60QHdVijUNZZJIL5fXRi+b3Rx/ggXn3G0pxBj1q7uE5zTXYVCXWOZ1sX3y1e6hFUSrUFTRbr4TFqf+xizHnUbyfVbFNQ1lml9sDuZdMpZ8okXnzF97jPah9VtNumeUqGusUzL4nPhXy8hZsknX3ymeLVz0zeh7lKge0qFusYyrYrPkfbkLPnki8/EmDk/ov2BusszuqVUqGss06L4HGlP0JLPA/HpjHbnxVx0cJnk4kkK6hrLtDTaX36kX1tHzpLPA/EZ0+d41aeDy2Qn85WgrrFMK+J7851+bB85Sz4fxGdMn9tVn65cJjy4jI/43n6hnzpA0JLPB/GlYfp0Ju0W3U8s1DWWiSu+t/YrJlUh5Gg+jQ/iM6bPafVqHdkpunKZgrrGMvHE986JY72CoFmnF+KjMBeXEZ7a0yA8spOB+N58cLfWM6Q57BLjhfhMmIvLCM/76gbSPQ3Ziu/l75/sH8R3h9/pbiLwQnxk+hzm9T1SN3hBdxMLdY1lIojv17dXDjLWa/DjF7qjCPwQ36nrlHZ1eeFp7CWoayzTQHz/u7x8e/WP0w2Wm4jJo9X4IT5KaXdWzUXX7Fyhe8mFusYy37/Uwv5B61H4jZ5UBp6Iz6S0T5JWrKNrSIg+I0VDXeMv/9GDCsET8RnTF/SSWGwzpC4u+UxaA3WNv/xJDyoEX8R3ok+qfejI3aAPKNqhW8mFusZbfgo5ly/EF/EV95Q+glFSi2V0wdxDupNcqGu8RcoBKSHeiM8clrlwj+Ril1l1bdEHFGmoa3zluzDD55H4Xug9FzeJffo0duEFXEpQ1/iKsBWfT+Ir7iiFOAnx7FQXFn0au4G6xlO+i3KwKzwS35nObnDh7NPVk9boNoKhrvEUcYbPJ/FROZdpUoxFBtV1hVerVlDX+MlXcYbPK/GRs2+AJGOPUXXZXbqJYKhr/OQtPaQgvBLf2aJSyXwXacYak+qyR3QTwVDXeImw4BaNV+IrHiiVBA9su9q1m094qXgFdY2XSDkKuhq/xFfc0Oqznd4woy4q+1xMDXWNj8hKZyA8E9/ZE60+y9WUdDaf9LqBJahrPETgbksJz8RHrnbLy755dU3x2Xwei+/nr/SEsvBNfMUjJRTLyz51RfFFO0tQ1/jHe3pAYXgnvuK6Vt8E6cYGXeqC0k8oUlDXeEdmYy0h/omPln2DpBwL6Iq5HgS4+Cq+H9ICqkP8Ex8t+yzmN/Sq60k/HkxBXeMboiqWVeOh+GjZN2dt00VHl8nPY/dVfPGK9nLCR/HRsu+RLfXpgzGln82noK7xiy8ivQwaL8VHy75ZS4U8dWin+HLVJahrvOK/l/RwAvFSfMULHeRpy+HgS2inj+L7Julshtv4Kb7iha6nFNy3oj4d2nlMV5YMdY1H/JCsPV/FVzw16psi/SRCh3bKr+DiofiERraE+Cq+4rF2OFhxtvsS2umd+H5KTGWowlvxUXpRMEYKSoAO7aSrioa6xhvEOvgIf8VH7r5gMvG6T13Fg/JJ3onvHT2WWDwWX3Ffiy+YSehx8Ca00y/x/RRYN+IWPouvWNDiC2aTedt1aKf8U1JKUNd4wXfZey0ar8VX3NXiC+YSxXnqI4p8CO30SXz/ifYxEH6Lr3hk9jwX+khIreBNaKdH4vssOK6lgufiKx4Yf1+SQtYQHztEVmy5i+/iK56aSLMELgctvqd0OdFQ14jnD3oe6XgvvuKFriIfBA+6SUxx0eLzIanBF/EJLRpxF//FVzzf0uIrGb/WPH46qQHi44PE+rg1yYH4imcmvy8IHvaQnmIB8XFDvofPkAfxFYv7ZtOzNeMH8XFD3mFgtcmH+Iov1rT2WjJ+Wnx7dCHRUNfIR95pYDXJifiKxZ3Q+I3HPcBvQv3Kh1xaf8TnienLjfgqxm8hpvy8SWT3R3yemL78iK/K+C2Mx4n2hPj44Yfpy5P4ii9Cp0MQTEaXnxbfIV1CNNQ1PuCF6cuV+IrFI4o2KxFZft6UcPFJfF6YvpyJr3heqMhvZiiS/nQJF4iPGT6YvryJ76b8gtnh5kFnEB9HfDB9+RPfLfkFD0eb6A/iY4kHEZ55FN9t+QWPxnsbeB8eqK/4UDnQL/H9lHo2UYV8ik/Jb0Xrrsz8zOhA7YR3byoH+iW+67/pqeSSV/GVeP6UUv2qeDAx3Hu73hLExxP5pi/H4ivxokDJfrd4MDMzc39U0wnxcUW86cu3+Epc7IdhZ7UZh/i4It705V58JU72tle10Gqx0Dan/gPxMUS66YP4iNPDwuqNLdAQWD62SDd9EF81J8eFzdVb26AQH1+Emz6IrxYvjo+PDwqm4DXExxfhpg/ia4AR36z63xf0kWioazziIz2ZTCC++pwZ8SG8rDGfL5vwH33RBaJPbID46nMB8UWhqfV5R190wT90D5FAfPXR4puD+JrQVHy/fKdvukCy6YP46qPF96jtPsTXkObrrr/omy6QbPogvvqQ+HQZiWf0mWioayzTXHyvftJXXSDY9EF89TmtEh8KKNUlwo7jP/RVFwg2fRBffY6V6mYhviZEEN9v9FUnyDV9EF99tPhmIL4mRBBf0aW3Qa7pg/jqQ+LT5eIhvrpEEZ9Lb4Nc0wfx1adafDv0mWioaywTRXxOvQ3/0k3EAfHV57BKfDilqC5RxOfU2yD2yDCIrz5HSnWTEF8TIonPqbdB6mmZEF99SHxDEF9DIonPqbdBqumD+OpD4tNnsu/SZ6KhrrFMNPE59TYINX0QX332qsS3Tp+JhrrGMtHE59TbINT0QXz10cm0YxBfEyKKz6m3Qabpg/jqo8U32tav/rNBn4mGusYyEcXn1Nsg0/RBfPUh8fWq/6zSZ6KhrrFMRPG59TaINH0QX32eKtUNtXWp/zymz0RDXWOZqOJ7Rd93g0TTB/HVZ12pbqCtQ/1nkT4TDXWNZaKKz623QeKRYRBffXQl3d62NvWfgD4TDXWNZSKL75J+4AaBp2VCfPXR5zh000kp5/ShZKhrLBNZfG69DQJNH8RXH32IUWdb20P1Xx8Kd1LXWCa6+P6gX7hBnumD+OqjNBe0eXQ0LXWNZaKLz623QZ7pg/jqcq40N1cS35T6gw9FXKhrLBNdfMW/6SduEGf6nIkvRpcwRddPelgS34T6gw/ZtNQ1lonR0269DeJMH8RXF8qlbfMnp4i6xjJxetqpt0Gc6YP46vJMaW6qJD6dU+RDWgN1jWXi9LRbb4M00wfx1UVnFE2UxOdNZDV1jWVi9bRTb8P1e7qLEJyJ74puIBcK7WwzwZ1b9KlkqGssE0t8br0Nwo4Mg/jqQqGdbW331B/W6FPJUNdYJpb43HobhJ2WCfHVhUI720xktQ/BndQ1lom3wHDrbZBl+iC+uoShnSa4c5k+lQx1jWXiic+tt0GW6XMmvj/oBnIJQzvb2ubVn87oY8FQ11gmnvgcextEmT5n4ntHN5BLGNpJkdUeBHdS11gmpvjcehtEmT6Iry5KcSq0k4I7T+ljwVDXWCam+Bx7G5KbviTFZv5H14iGM/Fd0g3EUg7tbDPHYx7Q54KhrrFMXPG59TYkN30eiE/0SfWKcmhnmzmnaI8+Fwx1jWXiiu+XH/RDNyQ2femJz9kMPF4zGHKgFKdCO9vaRtQfn9LngqGusUxc8Tn2NsRvzy08EB9dXy66ZK6KLmtrG1B/3KTPBUNdY5nYg93ZbItIOOmSL74fdH256ACXES0+HeLiQf0y6hvLxLc0br0NSU/LTE98b+lXtvlG15fLplKcCnBpM/XLPPCyU99YJr743tAvXZHM9CURX7xtxiR3aoT8RPYnSnH3tPja5tSfT+gv5EJ9Y5kW1lhuvQ0JTZ988ck9KDtkWQmuw4jvgfqz/Cou1DeWaUF8f9JPXZHI9MkXXws9wosTpbd5oz1TxUW+r4H6xjItdLVjb0Oyf/qTSCJe3ez39CvbiI+r1kUkHpD4dCEJ+bns1DeWaeXfWcfehkSmL4n44sV1faBf2UZg/eCb6Dx2VURC4YmvgfrGMq2Iz7W3IYnpS098rnZ9hR7UW2FX6W2MxNej/o98XwP1jWVaWmE49jYkMX1JxBcvl8fVvtNrur5YNpTeBkl8ner/LNHfyIX6xjItic+1t+Ffuk8LJBFfvNWWq6x+gWfG3GRN6U2n0ip0Rp94XwP1jWVaEp9rb0OCqVcS8f1F14jEL/Qj23yl68tlSclNZ/MpZtX/e05/JRbqHMu0Jj7X3obWT8tMIr5PdI1IuFr4infz6YSiBZIe+RrEF62mzrFMa+Jz7W1o3fQlEV8sybsK7YxlfjlyqtRmEooUY+r/ivc1UOdYpjXxOfc2fG115ZNEfLFCml0Zf/EVXA6V2qZJelS0eoP+TizUOZZpUXyuvQ0te7uSiO86TjLhR/qNbX6j64tlR6ltnKRHdXPFl+6kzrFMi+Jz7m34/pJuFJNE4osz8L/Rb2zT4nPzQRftHCbp+eJroM6xTKvic5VQU6bFghKJxBfDy/6SfmKb73R9ueiinX0kvRIL6v9LPxqaescyrYqv+JUu4IoWC0okEl+MzQ5X//gkcHEyYUWJTRftNGhfg/S8Buody7QsPtfehngb/2USiS/Gdudf9BPbfKDri0V7GoJ2Ul6JafX/9+lvpUK9Y5mWxffLT7qCM36nO8UikfhiLLi+0C9sI75op85pqHgaqIaS9GPCqHcs07L4nHsbrr+3MvFMJr7IQ/+Vq396xJcu21daq3ga2tr61AfSQ6updyzTuvhe0xXc0crEM5n4Ir8NV5Nu+fsterPTVE8y6O1O4ec1OPKstS4+596GliaeycQXeey7Cm6VX0NCH5JStdlJZVxkl4znJz7n3oZWnH3JxBc1rO03+rp1xGfSnimlVcKqFbpkvOxKEo76O4H4nHsbrq//jR1lllB8Efc7XWWxy49v0ZGd5pyGEB3duU1/LxNHq4wk4nPubWhh2ZdQfNFM3ytXceU/xSfz6WrV90l2hn71kewAM0f/2H6my7eCe29D/CD/pOKLZPo+0ZetI79m57ZSWlhDwtCtPpJdOPdf6h/LJOpu596GEjEdX0nFF8X0/U5ftY/4fCKTxt5PsiN0MvsL+oJEXG3sJypO7t7bUCKe+hKL70fTAjKvXBWQSJLBzwVdMLcquEyhC+ce0hck4szIJCpR697bUCJWaZXE4rv+1szP7e6hf4pPadD7LWHB3JBx9aHgc8JeOvvX9j3doSXcexsUceIdk4vv+ltDCbx0+A9OkgU4D3QmbVgwN0Tn027RNwTymbrHPq0mzhncexsU/0Uv5G5BfNdf39DFavCbqzw+hfgsdnM6WCWT1qDPCZOb0udyZ6O17AEiBW+D5p+oMY82xFd6J/XCSv90usGb9Fje7NlSOgtrdoa065S+C/qKMH5za18+JrB9aXgbND8/RduMsCO+6x9XNdT++sptR8jP5TNlA+l0sAo6pe8ZfUUSr/905GSoUHOkRSQNbwPx/cOb5j5oS+Ir8fXvG+Em//vTdbFS+UekFC+UyhaqkvkMk+rjAn1HAr9dXr65+vuLyyVGFV8/XF3GOyGL+JUukBLfPl/9fnl5WX9+Zk98ih9fvny6uvr0xVX23k3EF4ovPlMqmyXJVRhWH/M+LcVVPayWiWQP3W0FReK2tbArvlSRfx50saBUNkmSq6ArmC3Sd3giU3zpeBvq4pH45Ie3mNPYK5XLQvTJ7LyLKMkUn7MaetHwSHwtTft5oYsn9ZDkqnioPme94yJUfGl5G2rjj/g8mHXq/ZbwNPZq9I4L6xgXoeJ7mZa3oSb+iC9RqBEP9Jm0d/db2toG1V+wruMiVHxpehvu4o34WqxTygpdv+V2fIuiS/0F60WfVPGl7G24iTfik1+9pVhcVBK7lU9keKT+hvOiT6r4MvU2eCO+BvGkUqi75BOw6BMrviy9Db6Iz4PtlvpLPgGLPrHiy9Lb4Iv4PNhuqb/kE7Dokyu+DL0NnojPh+2WBks+/os+ueLL0Nvgifh82G5psOSjRR/j06Hlii9Db4Mn4vNgu6VeVLVBL/pW6ZsMESy+7LwNfogvxtlkfKmZxR6iF33LfE9sECy+7LwNfohPftWyEvqUhjpLPlr08T0jU7L4MvM2eCE+LwyfORWzzpKPfUKtZPEV3dWybIwX4vPC8DVc8rFf9IkWX1beBh/E58NWZ5MlH/tFn2jxOTuqtQk+iC9R3WI2NF7ycV/0iRafu1OzGuOB+PwwfE2WfNwXfbLFl5G3wQPx+WH4miz5uC/6ZIuvmE6Fr9vIF58fhq/Zki8M7zyhr3NDuPiy8TaIF9/PeG+ZLY0COw160XdEX+eGcPFl420QLz4f0hlK6OOJFhos+eh4aK7VO6WLLxNvg3Tx/Sf+IGiDrtg5QzKrja7eydXZIF18mXgbpIvvRkl6wWhHwxDJrA5z6jtM04qkiy8Tb4Nw8cU5epAzJp2oi1RWB+1sWKdfMEO8+LLwNsgWX7LDERmxr3R1+1DM2/SrL63QL5ghXnxZeBtki+93arV4VpWuRklk9TDn9PEMcpEvvgy8DaLFJ/8YaOJ8Wcnqzrl8t5lW3+KZzi5ffBl4GySLr/Hh75LQdcsekcTqo4NcntBveOGB+NL3NggW3w/55/GFbChVNQpvMZjTil7Qj1jhgfjS9zYIFp8XWXyaM30adC9JrAEz6ns79CtWeCC+9L0NcsXnSWiL4kBpap4E1ogR9cU1+hUrfBBf6t4GseL7RA32gW2lqbsH0t6Fb3C1D+Ir/ku/TQup4vvXk7AyjQ6qHiCBNUSfkskxuNoL8aXtbRAqvm8+VKgOea4UtdBO+mrIqPrqFv2OE16I75eUvQ0yxffdjwRaQgdV3yd5NaZHfZVjcLUX4iu+px+nhEjxfW3pzbLliVLUIMmrCVzPbPBDfCl7GySK74s3znWNCaruJHU1YUJ9l2FwtR/iS/kxBIrvk097LSUiBVWH9KkvL9EvGeGJ+NL1NsgT39/UUm/QQdUjJK5mtM+rbx/QT/ngifjS9TaIE98f1FBvOFFqCrpJXE3RSX0b9Fs++CK+VL0NwsT3w5+YspAdpaZGNQNvoued/PY7fRFfqt4GWeL77JN7j9B7nVFnnSV0MQl2fnZfxJeqt+G2+Iq/fczumNwmfH9HbfQJXbYs6l6nQhcxY1c81xvxpeltuCO+YvHlu7Rj3KLx0S8PA6GL5UbzsBvuqR+wi+/0RnxpPkkN8ZV4/eEH/T0bvvlw8vNdzlaUlCJ62A06vpNbXpE/4kvR21BbfKWFJzPz95dnzr0QnU0ULa4zROcVPabfc8Ef8aXobagnvhKv/sjm+Iga/ONVLGc1Ooc9SjZRBZNXdEoXYIJH4kvP29BAfCVY6O/nh9ZfJHdM5aQ+klVEdD47szpK778wo/V98V8+0yWc03QD8dW7f0gE2fDjykP3Qpk9paM5ElVUdB0lpvU7gWVevvuY1ZHx3/7wdK1H6NCyMRJVVDp0/U5+IWbAEa///Cd9/9+/3hTFrYNJaGhar/M2U+pXTOvGAzdc/pXmDujnP/xd6oXoNNrooWUhum481/OKgCtevrn6koIF/PnpnZce9dvEDS0j2nmGmIEUeP3ug0sT+OPjW78XemV08ZY4oWUhOqWWYykXkAqX7z85UOC3f658OXEvArpkYJzQshBdyoXt+ewgFV6/ubImwe+fry5zMdcsYwpVxwotC9GlXFiWrgbpUpLgX1+SLAW/fr5647M3rw7PlILihZaF6BKC3ELMQIb8evnu6sOXL99IUk3578vHq/eXlzlZ4N1lUykoXmhZSLf6afCcLgRAFb9dXr69UnykSBpCf3Z19e7y8pK+mV9M/YiYoWUhD9RvN+lKAIBY6Ey+uKFlIcbVhy0XAFrAZPINk5jiYlx9BboWACAG+jTahRacfAad1beCKBcA4vNYqWeCpBSfTh1dfUgXAwBExkS3xI6prqCjq+FtACA22s8QtUh8LUwhJWYJ7QDw50SnsPeTkFpiVl2BX+1qAJizq5Qz11J0S8iAugS8DQDEw/gZ4icTVQNvAwAtkNDPYNABnvA2ABAL7WdoLayzgvE2MDymFgC+mAMaEvgZDNrbsEbXBABEQJfKTeJnMMDbAEBMLPgZDNrbgDJmAERGFy1L5mcw6PK58DYAEBXjZxglASXBeBtQTgKAiBwqxST1Mxj0SZnwNgAQEe1nmCL5JMN4G/bpwgCAhui6Scn9DIZpdS2YPgAioQ1f/BrxtTHeBpg+ACJgDJ8FP4MBpg+AqFg1fDB9AETGRJZZM3wwfQBERWew2zN8oelDMRcAmmDd8JHpQzEXAJpg3fCFpg+ZRQA0xIHhg+kDIAoODB9MHwARcGL4YPoAaI42fI9IMvaA6QOgCcbwtXo2SgNg+gBojDZ8NpJobwPTB0BDLrREHBi+trb76sowfQDUwZyG6cDw0VmZMH0A1MaUTXJi+KiUEkwfADVxtuJTGNOHCE8AamAO5HNk+Mj0IbkBgBroPD5Xhi/c8EReHwB3MAnszgwf+fqWUMMTgFucLSptzDozfG1t3bqQ2VO6HwCA2FfKsB/VWc24usPyBd0QAKA510WqZ0gmbuicV/fYpDsCADTav26rVmc9hvVNcGgRAFUY//o0icQV5uQGeNoBqEL71+2cztCIAXUbBJkBUMGkEo2TRByiPe1P4GkHIGRVaWLOueFra+tTN4KnHYAQ5/71CtrTjiAzAAjHgWXVmCAzeNoB0KTgX68woe6Fg6IBUJhDoG2XC6yHOS1zne4NQK4pKDUEvSQO5+iDouFpByAt/3oFeNoBINaVFlwHllVjgszgbgC551hLIQX/epl27WnHngvIO2dPlBLmU/CvV+hRt0R2A8g7ZrdliGSREtrdgBBPkG9e6N2WBySKtOjQey6LiHMBeUYHdS6kuNtiMNkNu9QIAHLIkRbBGEkiRXT1+GU4+0BuOdGxLY9SCeq8SZeOc1mjdgCQO4yLr48EkSoj+tZ71BAAcoZx8U2RHNLFOPtQxRPkkyxcfBWMs2+D2gJArsjExVfBOPsOqDEA5IhsXHwV4OwDuSUjF18F4+wrUHMAyA2ZufgqGGffC2oQADkhOxdfBePsW6UWAZATttW4z8bFVwHOPpBDsnTxVaDMPkw8QY4414fxZeXiq9CrmhGsYccT5AcTV5aZi6+CqaaEHU+QG0yF6vskgCwxE8/gOTUMAM8xO53ZTzoV5qxoHJ0CcsKWGu/BAA3/jDE7ntvUNAC8Zk8P90ka/Jkzo5uDGE+QA0xM56MOGvuZ06VPal9BchHwnzU11tOrDt+cQd0gVBIE3mMSidIsktuUKd2kI2ogAJ7yXE86H2Ya03mbTp1ctHRBTQTAS0z2epaJRLUwh0Ujwhp4jYmnHqFBzwaT1b5DjQTAQw70IM8ue70e7Q9Vu1DHE/jLuQ5tWeiiIc8IU04JgS7AWzb0EB+kAc+KUd00BLoAT9nXAzy9M2jj0P5ANw7+BuAlxsswxyKe+i4mwhqJtcBHTAJt0E+DnR1DunlPzqm5APjDph7crEJbbjKpG4gwM+AdJqzsAavQlpsYf0OwTw0GwBNMxaR5hl6GCrTsQ1o78AqTvM53wWfo141cRHYR8Igzk0eUaX3qKIzrZiLIE3jEUz2oZ2iI84W8fahmBrzBVCvj6uGrxqS1B8fUcACEc7GkRzSj5PX6mOwi5PYBPzA5fMEwDW/mmCBPFLEGXmDKU/MM6ayBqWb2lBoPgGDMQXx8qpU1wxSVCA6p+QCI5VSHUy/00NAWQA9CrIEXUDg1gzNRojOsm4wQayAcE07Npjx1NKZ1o7foEQAQifGu86oU2JyOR7rZyGsHgjHHMix006AWwz297EOCA5DLsd5s4R5OXQvja8fxKUAqL0xkC7sqnVEwmy5LKCYIREIbncI2W0JMHV2kFwGJnK3q4TsjbLOlzH3dfMSZAYGYqDI5kS236TBVJTbocQAQw44eurzrRjSmy8SZ7dIDASAEk8K3ICKNqB4mzgxRnkAWpkAuz8rw0RnQD4GKSkASF6ZeEvuaLc0Y0Y+B1Fogh7PHetBO0RAWjCmkixhrIIYtPWQ5F8iNSrtJrUWMNRCCiaZ+JKBeUnMoxnqdHg0A1pho6nlx0dS1oXpmKCsBBGDKRgR9NHjF02scDqjlCdhjHHyyUtcbM2ieCPlFgDkHxsHH+CSw+Jhqgji1FvCGnOtCMxnqMaYfKnhGDwkAQ05NBt99D5wMNzDuvmXk1gK2UF1477TX1jalHwyBZoArFyZ7dlZsFlF92k12HzLbAU8oc91H7ZUPD0OYJ+DIuQnofCQ4g68RHbP68RahPsAOOnzWV+2VA80WEWQNmHFmgqnn79FQ9ZAuo77HUB/ghSkL77P22tq6TZgnaioBVphqSQveBHTW5p5R3xbUB/hgkoh81145yHqTHhuAzCnoISm9YksU+oz6kN4HmJAf7ZVTHDYw8wQcIO2N0vD0HFIf1n2AAfnSHtQH+JA37UF9gAvbZiTmSHtl9a3C2w6yxPj38qW9svoQ6wIyJJ/aK3scoD6QGXnVHtQHsoa0J/LY56SE6kOGEcgAymPIh2/9LqQ+5PeB9Mm59spR1lAfSJvcaw/qAxkB7ZUI1YeqSiBFzqE9BakP9TxBelCNwIUBGoS5pdtUlgj26MUA4Jjnpjau/7mzzaG6LsEuvRoAnPLMnMcA7Sm6HuqXEawjzBq4x5x9GcxDe5oOc3B0sIVgF+AaU64lmPO6Tlkc2s05Dgh2AY452zAjbdbb2rgtMG7eCVwOwCXnpix1MOPleQwtM2TeytIxvSYArHPxxIyySf/OAEtGvwn0XD6kFwWAZU5X9BALxmjIgTI9xt0e7NCrAsAqdNx63sNaahM6/LbpZQFgEbgYGkJHiAWbcPgB2+yasQUXQz3ap80bWoPDD1jljLLW4WJowIR5R09e0EsDwAIncDFEYcS8paVn9NoASMxz2uaEi6EJA8blEBToxQGQkD3a5oSLoSm95HLYxMIPWOCMilLDxRCFbspywMIPJCdc7s330vACDemgTU8s/EBSwuUetjkjM2reGKJdQDKOaLk3ha2W6PTTwg8naIIEUPJeMEzDCkQiXPghxQ+0yvmqGUNY7sWlvPBDkhFoiVNTowzLvVYIF3779DIBiMEhlntJwMIPtAyWewnBwg+0BpZ7ySkv/JDfDmLwHMs9G4QLPxT1BJHZoUGD5V5CwoXfExQ2A5E4oSknlnvJCQ9zWMauJ4jAgTmJAcs9K3RMmreJitagKWfhLueDTho+IBmDNPVcgcMdNOTFYzNSglEaOiAx3Q/oneIoI9CAMI56rocGDrBA+5h5q8EaXH6gDud0EEMwjVItdumbMy8WLj9Qm9C5tzBEQwZYo/O+ebdw+YFahM69h900YIBNRqi2Elx+4DZl594EHOtu6KFYT7j8wE3Kzr1+GirAOhWX3wm9dQCK51SPGs49t4Quv6UjevEg9xxQkSQ491xTdvltobIgKFE2e3Duuafs8ltGbTNQMXtw7qVCL+27BGswfjmnbPbmUY86JWD8gOaINjmD+9hpSY/Q6QDjl2NOtmgQwOylS5XxQ8BLPimbvUms9tKmbPyePKfeADmibPbmcMh6BpSNX7AL45c3YPaypuzzg/HLFy/CSE6YvQwJY62DpygxkRvOdihnFmYvW8rGb2mPugZ4zsET6vJHMHtZUzZ+a5h75oCLcKMlGEPyUPaUjV+wjbmn55wVwhnnQ0Ry8mCEUh2CJST6ec2zMJATZo8PnWGeX/AYc09veVGecc6gVAQnemapX4J15Nl6yfluOOOcQ7o6N4Yx9/SZw3DGuYAZJ0M6Jqh7gieobO0Zp6FXPbiPU7940lPe99zA3NMjzsMDGODa48wQldYNlpDp5w1H5RnnCGacnOkYp44KFp9R3wHRHIdnnwTTyJjlzr3y3PPxAfUfEEtlsfcQ5+1JYDCcewar8PqJ5mKTOjKYH6HOBczpKGf6BZsoMyGWk7A6UhBMYsYph+5p6rUgWMehYiI5fxo61YMZxHHKoneGei5Yfgq/gzjOCmGqejCLgBZ59JcjzpYKSHeQxX45gvoRKpPJZOAR9WCwtI8yL3I4pGMug2BuGJ49sQyXNz5XcLKKEA7Kjr35UZSJkEz7aBhvHTzBgdICOC479hbGscUpnY6xsM5EsAjrx5xnZenBu+AHneV0h2AFaz/GHJUnnME00mV9oWuK+hQ7n2w5Oypvs8Cx5xf37lO/luQHvx8/zivOhWAWeUO+UfG6B8uIeuHFecWlDp+6n9yrxJwF66fU7yBzTiqBZMEMpOcr3eUqZ0GwiVoTLLiohE8H95E25DNd42XHQ7CKfL/MOd2gzigxfY86CfhK51jZ7Y5024x5Xq7EGQSTqIyUBzpGy0FnwZM9eB6y4nCNOkFFs0B6eaG9EvMZLG0j3zYDTgoV38L8GKJZcsVgOeMhCLYQ9Zkyz6uWenMjCJ/OHQPlfL8gWNyB4z01zo7CE/ZKPELSUD7pr4S9BMvrKLaUChdPKw71YBapsvmla6Ky9RmsHSHq2jUHVfubwSRCOPNN+9BDGgolVnYRd+aQ8/1K7HRpqYddFtDWW8l5CIINuP4c8WK9EkUWzAzQywd5p3Os4noIFgswf9Y52avk6gULE8jWA1UMlqvMl1iD690mZ4flytMlHg3BtQBucW+yEvZZmn7inBVLPN+u2t4MppGsB2rRMVLleQ9WtpF2lJiLQtUeSzA/higyUJdq118QPNnB8i8B53uV6M0SDwbhTwcN6Ryu8j0EwRacfy3yrCqErLTSG8MmC4hA93jV5mewvA7vQ2xOtytx06Xp5iSyZEFk+m7svqw8Rdp7DE4LVcGbQTA9gOkmiEX74I3l39L6IeafUTjert5iCWaHEMgCWuDW8m958wi5Dw05O1yvditgoQeScHP5FwSr+9j/rMPJ0WZV/BgWesACN5d/QfCkAP/fHS72b3gVsNADlugYmKrKPCqxuI0NmCpOd29ssJSUh4UesEj/xM35JzZgiFsbLMH8FGwesE7P+I39lyBYK+TcAJ7ub91Y5gVzE6g4DRzRPVJV+EWxvLWT0xXgi72NGzubQfBwHFVvgVO6hionrhiWNvdyVn3w5Gj95lwzCGaH4VUAKdAxOHVzAzQIFtfz4gM8f7Z9a3slCO5jgwWkR3v/ZHX6kebJ9jPPc3DPDnarEtINc1MDSI4FadM9NHVzB7TE4+0jT9eAF4dPKwemEwvTw1jmgay4Nzx9ewYaLK8WDryag54dF7aqkxQMMyOo/Aeypmfk9hZMicWNHS/8EKd763eWeEHwYAyhY4ALvWPVBZhC1p4eCd4HPXm2e2eiWWJ2vB9edMCL9v7xW05AzdJW4UBeLMzxzubdiWYQPJoYwL4m4Eln3+jMnTVgiZXV3f1jEVuhZ8dHhdXbTjzNzPgAyh8B5twbmqhlAktGcPXpzjHbrZjz473d1VrmrmTwpoaxuQKk0N47ev+OH8KwvLpdOGaVFXhyvPN09VasWMjCzGgf/HhAHF0D47W2YQyr64XDrDV4cvyssL16Mza6itmJQbjxgGB6hqduJUPcYGV1s7BznLJX4vnxfmGj9sqOeDQ92os9TeADPf2j07XXgWXWVp8WnrldEZ4fHxR2V2u5D6p4eH90EE484BvdfaNTNfzxN1laXS0tCguF4+NjCz7Ci9JldgqFp6WL1t5LqWJ2erQf+yrAZ7p6Rydq+iNq8nh1daskxYKam4aQsG7wnP7uuDSjVGyU5EaXaM7M1GgfMoJAXujsHRmdnqmzIZoWczPToyO98N6BfNLVOzA6PlN/U9QJD2bGRwcgOgAMve5NoTF02E0BoA7dvb29o6OldeHMzM2Cha1QushU6WL9vb3YSgEgHu0lKQ6W5FOam4bUnKPO0l/OzEyqbw+XfoYQaNa0tf0frw7txnWj1RcAAAAASUVORK5CYII='
sg.theme('DarkGrey6')
menu_def = ['&About', ['&Automation tool for Agile']],['&Help',['&Reach out to rjameel@its.jnj.com']]
layout = [  [sg.Menu(menu_def, background_color='lightsteelblue',text_color='White')],
            [sg.Text('Enter your Agile user name'), sg.InputText(key='usr')],
            [sg.Text('Enter you Agile password'), sg.InputText(key='password', password_char = '*')],
            [sg.Text('Enter download location')],
            [sg.Input(), sg.FolderBrowse(key="dl")],
            [sg.Text('Select the chromedriver exe file')],
            [sg.Input(), sg.FileBrowse(key="driver")],
            [sg.Text('Select the excel file consisting Label/COs name')],
            [sg.Input(), sg.FileBrowse(key="exlo")],
            [sg.Text('Enter the type of Automation you want')],
            [sg.Combo(['Banner Copy Download', 'Data Scrubbing', 'Native file Download', 'CO Data Extraction', 'CO Status Extraction', 'CO Affected Items','Banner Copy Validation'],default_value='Banner Copy Download',key='fac')],
            #[sg.Listbox(values=['Banner Copy Download', 'Data Scrubbing', 'Native file Download', 'CO Attachments download'], select_mode='extended', key='fac', size=(30, 6))],
            [sg.Button('Ok'), sg.Button('Cancel')] ]
window = sg.Window('Agile Automation (Developed by rjameel@its.jnj.com)', layout, icon=logo)
event, values = window.read()
if (values[0] == "Reach out to rjameel@its.jnj.com"):
    import win32com.client as win32
    window.close()
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = "rjameel@its.jnj.com"
    mail.Display(True)
    exit()
if (values[0] == "Automation tool for Agile"):
    text = "This is a Web Automation tool which can automate tasks on www.Agile.com as per user demand. There is a drop down menu which allow thw user to select what kind of automation he/she wants"
    sg.popup_scrolled(text, keep_on_top=True, title="About", font=("Calibri", 12), size=(50,10), icon = logo)
    #sg.popup_no_buttons('This is a Web Automation tool which can automate tasks on www.Agile.com as per user demand. There is a drop down menu which allow thw user to select what kind of automation he/she wants', non_blocking=True)
if event == 'Cancel' or event == sg.WIN_CLOSED:
    window.close()
    exit()
window.close()
wb = openpyxl.Workbook()    
ws = wb.active
try:
    Filename = pd.read_excel(values["exlo"])
except:
    layout1 = [  [sg.Text('Unable to open the selected excel. Please check the file')],
                 [sg.Button('Ok')] ]
    window1 = sg.Window('File Error', layout1)
    newevent,newvalues = window1.read()
    window1.close()
    exit()
l=Filename['File_Name'].tolist()
listToStr = str(values['fac'])
#listToStr = ' '.join(map(str, values["fac"]))
wb = openpyxl.Workbook()
ws = wb.active
#inst = webdriver.Chrome(r"D:\Users\RJameel\OneDrive - JNJ\Desktop\Download Location\chromedriver-win64\chromedriver.exe")
#inst = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
inst = webdriver.Chrome(values['driver'])
inst.get('https://agileprod.jnj.com/Agile/default/login-cms.jsp')
window_before = inst.window_handles[0]
usr=inst.find_element('id','j_username')
usr.send_keys(values['usr'])
pas=inst.find_element('id','j_password')
pas.send_keys(values['password'])
inst.find_element('id','login').click()
inst.implicitly_wait(5)
counter = 0
for i in l:
    error_counter = 0
    err=[]
    try:
         window_after = inst.window_handles[1]
    except:
        layout1 = [  [sg.Text('Wrong Credentials. Try again')],
                    [sg.Button('Ok')] ]
        window1 = sg.Window('Credential Error', layout1)
        newevent,newvalues = window1.read()
        window1.close()
        inst.close()
        exit()
    inst.switch_to.window(window_after)
    inst.implicitly_wait(2)
    try:
        WebDriverWait(inst, 20).until(EC.element_to_be_clickable((By.ID, 'QUICKSEARCH_STRING')))
    except:
        layout1 = [  [sg.Text('Unable to load website. Please rerun')],
                     [sg.Button('Ok')] ]
        window1 = sg.Window('Automatic Scripts', layout1)
        newevent,newvalues = window1.read()
        window1.close()
        inst.quit()
    inst.find_element('id','QUICKSEARCH_STRING').clear()
    search=inst.find_element('id','QUICKSEARCH_STRING')
    search.send_keys(i)
    if listToStr == "Banner Copy Download" or listToStr == "Data Scrubbing" or listToStr == "Native file Download" or listToStr == "Banner Copy Validation":
        inst.find_element(By.ID, "selector_elm").click()
        inst.find_element(By.LINK_TEXT, "Items").click()
    inst.find_element(By.CSS_SELECTOR, ".quick_search").click()
    try:
        if listToStr == "Banner Copy Download" or listToStr == "Data Scrubbing" or listToStr == "Native file Download" or listToStr == "Banner Copy Validation":
            WebDriverWait(inst, 5).until(EC.element_to_be_clickable((By.ID, 'col_1001')))
        else:
            WebDriverWait(inst, 10).until(EC.element_to_be_clickable((By.ID, 'col_1047')))
    except:
        try:
            inst.find_element(By.LINK_TEXT,i).click()
            time.sleep(3)
        except:
            err.append(i)
            err.append("Label not found")
            ws.append(err)
            wb.save('Output.xlsx')
            error_counter = 1
            pass
    if listToStr == "Banner Copy Download" and error_counter == 0:
        try:
            Filename = pd.read_excel(values["exlo"])
        except:
            layout1 = [  [sg.Text('Unable to open the selected excel. Please check the file')],
                        [sg.Button('Ok')] ]
            window1 = sg.Window('File Error', layout1)
            newevent,newvalues = window1.read()
            window1.close()
            inst.quit()
            exit()
        bcd(i,values["dl"])
    elif listToStr == "Data Scrubbing" and error_counter == 0:
        if counter == 0:
            header=[]
            header.append("Number")
            header.append("Document Type")
            header.append("Lifecycle Phase")
            header.append("Description")
            header.append("Class /Document Category")
            header.append("Document Security")
            header.append("Product Line(s)")
            header.append("Rev Incorp Date")
            header.append("Rev Release Date")
            header.append("EFFECTIVE FROM")
            header.append("Product Family")
            header.append("Material Group")
            header.append("Material Type")
            header.append("Responsible Location")
            header.append("Plant")
            header.append("Language")
            header.append("Formula")
            header.append("Transfer to LMS?")
            header.append("Service Library")
            header.append("Point(s) of Use")
            header.append("Notes")
            header.append("Create User")
            header.append("Document Owner")
            ws.append(header)
            wb.save('Output.xlsx')
            counter = 1
        a=scrap()
        b=[]
        b=a.scr()
        ws.append(b)
        wb.save('Ouput.xlsx')
    elif listToStr == "Native file Download" and error_counter == 0: 
        nfd(i,values["dl"])
    elif listToStr == "CO Status Extraction" and error_counter == 0:
        if counter == 0:
            header=[]
            header.append("CO Number")
            header.append("Status")
            header.append("Date Originated")
            header.append("Date Released")
            header.append("Implemented Date")
            header.append("Approval Date")
            # header.append("Product Line(s)")
            # header.append("Rev Incorp Date")
            # header.append("Rev Release Date")
            ws.append(header)
            wb.save('Output.xlsx')
            counter = 1
        a=cse()
        b=[]
        b=a.scr()
        ws.append(b)
        wb.save('Ouput.xlsx')
    elif listToStr == "CO Data Extraction" and error_counter == 0:
        if counter == 0:
            header=[]
            header.append("Number")
            header.append("Status")
            header.append("Change Type")
            header.append("Change Category")
            header.append("Description of Change")
            header.append("Reason for Change")
            header.append("Reason Code")
            header.append("Product Line(s)")
            header.append("Workflow")
            header.append("Change Analyst")
            header.append("Originator")
            header.append("Date Originated")
            header.append("Date Released")
            header.append("Final Completion Date")
            header.append("Product Family")
            header.append("Site(s) Approving Change")
            header.append("Risk Assessment Required?")
            header.append("Risk Assessment Summary")
            header.append("Risk Assessment Details")
            header.append("Regulatory Action Status")
            header.append("Regulatory Assessment Summary")
            header.append("Regulatory Assessment Details")
            header.append("Qualification / Validation Required?")
            header.append("Qualification / Validation Details")
            header.append("EHS Review Required?")
            header.append("EHS Review Summary")
            header.append("EHS Review Details")
            header.append("Project Number")
            header.append("Reference Number")
            header.append("Description Change Details")
            header.append("Clinical Research Approver(s)")
            header.append("Environmental Health/Safety Approver(s)")
            header.append("Legal Approver(s)")
            header.append("Manufacturing Approver(s)")
            header.append("Marketing Approver(s)")
            header.append("Operations Approver(s)")
            header.append("Quality Approver(s)")
            header.append("R&D Approver(s)")
            header.append("Regulatory Approver(s)")
            header.append("Service Approver(s)")
            header.append("Supply Chain Approver(s)")
            header.append("Additional Information")
            header.append("TVV/LMS Training Updates")
            ws.append(header)
            #wb.save('Output.xlsx')
            counter = 1
        a=co()
        b=[]
        b=a.scr()
        ws.append(b)
        wb.save('Ouput.xlsx')
    elif listToStr == "CO Affected Items" and error_counter == 0:
        inst.find_element(By.LINK_TEXT, "Affected Items").click()
        inst.find_element(By.ID, "MSG_More_1span").click()
        inst.find_element(By.LINK_TEXT, "Full Display Mode").click()
        WebDriverWait(inst, 5).until(EC.element_to_be_clickable((By.XPATH, """//table[@id="CHANGETABLE_AFFECTEDITEMS"]/tbody/tr[2]/td[2]/div/div/table/tbody/tr[2]/td[4]""")))
        time.sleep(3)
        f=[]
        for rj in range(2,40):
            try:
                #//table[@id='CHANGETABLE_AFFECTEDITEMS']/tbody/tr[2]/td[2]/div/div/table/tbody/tr[2]/td[4]
                status_path = str(inst.find_element(By.XPATH,"//[@id='CHANGETABLE_AFFECTEDITEMS']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["+str(rj)+"]/td[4]"),text)
                print(status_path)
                descp = str(inst.find_element(By.XPATH,"//table[@id='CHANGETABLE_AFFECTEDITEMS']/tbody/tr[2]/td[2]/div/div/table/tbody/tr["""+str(rj)+"]/td[8]"),text)
            except:
                break
            f.append(status_path)
            f.append(descp)
            ws.append(f)
            wb.save("CO Affected Items")
    elif listToStr == "Banner Copy Validation" and error_counter == 0:
        if counter == 0:
            header=[]
            header.append("Document Name")
            header.append("Banner Copy Name")
            header.append("Description")
            header.append("Type")
            header.append("Size")
            ws.append(header)
            wb.save('Output.xlsx')
            counter = 1
        f=[]
        cntr = True
        banerr=[]
        WebDriverWait(inst, 20).until(EC.element_to_be_clickable((By.ID, 'Actionsspan')))
        inst.find_element(By.ID,'Actionsspan').click()
        inst.find_element(By.LINK_TEXT, "Banner View").click()

        inst.implicitly_wait(2)
        inst.switch_to.window(inst.window_handles[2])
        inst.implicitly_wait(2)
        j = 2
        while(cntr):
            f=[]
            try:
                WebDriverWait(inst, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/table/tbody/tr["+str(j)+"]/td[2]/span/a")))
                f.append(i)
                f.append(str(inst.find_element(By.XPATH, "/html/body/div/table/tbody/tr["+str(j)+"]/td[2]/span/a").text))
                f.append(str(inst.find_element(By.XPATH, "/html/body/div/table/tbody/tr["+str(j)+"]/td[3]/span").text))
                f.append(str(inst.find_element(By.XPATH, "/html/body/div/table/tbody/tr["+str(j)+"]/td[4]/span").text))
                f.append(str(inst.find_element(By.XPATH, "/html/body/div/table/tbody/tr["+str(j)+"]/td[5]/span").text))
                ws.append(f)
                wb.save('Output.xlsx')
                print(f)
            except:
                if(j == 2):
                    f=[]
                    f.append(i)
                    f.append("No banner copies found")
                    ws.append(f)
                    wb.save('Output.xlsx')
                cntr = False
            j = j+1
        inst.implicitly_wait(2)
        inst.close()
        inst.switch_to.window(inst.window_handles[1])
        #"""//*[@id="CHANGETABLE_SIGNOFF_HISTORY"]/tbody/tr[2]/td[2]/div/div/table/tbody/tr["""+str(i)+"]/td[4]").text)
inst.implicitly_wait(5)
layout1 = [  [sg.Text('Program Complete')],
            [sg.Button('Ok')] ]
window1 = sg.Window('Automatic Scripts', layout1)
newevent,newvalues = window1.read()
window1.close()
if newevent == "Ok" or  newevent == sg.WIN_CLOSED:
    inst.quit()
    exit()


