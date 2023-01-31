import barcode
import PySimpleGUI as sg
from barcode.writer import ImageWriter
import openpyxl
from barcode import EAN13
layout = [[sg.Text('Enter the barcode height'), sg.InputText()],
            [sg.Text('Enter location for the file to be generated')],
            [sg.Input(), sg.FolderBrowse(key="dl")],
            [sg.Text('Enter excel file consisting the barcode info and file name')],
            [sg.Input(), sg.FileBrowse(key="exlo")],
            [sg.Checkbox('Check if text to be printed along barcode', default=False, key="-IN-")],
            [sg.Text('Select the type of barcode : ')],
            [sg.Listbox(values=['Code39', 'EAN13', 'Code128'], select_mode='extended', key='fac', size=(30, 6))],
            [sg.Button('Ok'), sg.Button('Cancel')]]
window = sg.Window('Barcode Generator', layout)
event, values = window.read()
window.close()
wb_obj = openpyxl.load_workbook(values["exlo"])
sheet_obj = wb_obj.active
listToStr = ' '.join(map(str, values["fac"]))
if event == "Ok":
    for i in range(1,sheet_obj.max_row+1):
        bdata = sheet_obj.cell(row = i, column=1)
        try:
            filename = sheet_obj.cell(row = i, column=2)
        except:
            filename = sheet_obj.cell(row = i, column=1)
        number = bdata.value
        barcode_writer = ImageWriter()
        if listToStr == "Code39":
            my_barcode = barcode.codex.Code39(number, barcode_writer,  add_checksum=False)
            pat = str(str(values["dl"]) + '\\' + str(filename.value))
            if values["-IN-"] == True:
                options = dict(module_height = float(values[0]), add_checksum=False)
            else:
                options = dict(module_height = float(values[0]), add_checksum=False, write_text=False)
            my_barcode.save(pat,options)
        elif listToStr == "EAN13":
            my_barcode = EAN13(int(number), writer=ImageWriter())
            options = dict(module_height = float(values[0]),add_checksum=False,write_text=False)
            my_barcode.save(pat,options)
        elif listToStr == "Code128":
            code = 'TORICODE'
            my_barcode = barcode.get('code128', code, barcode_writer)
            options = dict(module_height = float(values[0]),add_checksum=False,write_text=False)
            my_barcode.save(pat,options)
elif event == "Cancel":
    exit()
