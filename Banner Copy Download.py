import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from pathlib import Path
import PySimpleGUI as sg
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time 
class bcd:
    def __init__(var,label,down):
        var.label = label
        var.down = down
        print(down)

        WebDriverWait(inst, 20).until(EC.element_to_be_clickable((By.ID, 'Actionsspan')))
        inst.find_element(By.ID,'Actionsspan').click()
        inst.find_element(By.LINK_TEXT, "Banner View").click()

        inst.implicitly_wait(2)
        inst.switch_to.window(inst.window_handles[2])
        inst.implicitly_wait(2)
        WebDriverWait(inst, 20).until(EC.element_to_be_clickable((By.ID, 'downloadLink')))
        inst.find_element(By.ID, "downloadLink").click()
        time.sleep(2)
        while (True):
            chrome_temp_file = sorted(Path(var.down).glob("*.crdownload"))
            #downloaded_files = sorted(Path(var.down).glob('*.*'))
            print(chrome_temp_file)
            #print(downloaded_files)
            if len(chrome_temp_file) == 0:
            #if (len(downloaded_files) >= 1):
                break
        inst.implicitly_wait(2)
        inst.close()
        inst.switch_to.window(inst.window_handles[1])
        
class scrap:
    def scr(self):
        '''wb = openpyxl.Workbook()    
        ws = wb.active'''
        f=[]
        try:
            f.append(inst.find_element(By.ID, "col_1001").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1081").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1084").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1002").text)
        except:
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1082").text)
        except:
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1068").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1004").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1017").text)
        except :
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1016").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_12089").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008074").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2023").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2029").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2092").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2090").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2091").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2007").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2024").text)
        except :
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2021").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_2000008063").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1080").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1420").text)
        except :
            f.append("NA")
            pass
        try:
            f.append(inst.find_element(By.ID, "col_1271").text)
        except :
            f.append("NA")
            pass
        print(f)
        return list(f)

class nfd():
    pass
        
sg.theme('Topanga')
layout = [  [sg.Text('Enter your Agile user name'), sg.InputText()],
            [sg.Text('Enter you Agile password'), sg.InputText()],
            [sg.Text('Enter download location')],
            [sg.Input(), sg.FolderBrowse(key="dl")],
            [sg.Text('Enter excel location of files along with extension')],
            [sg.Input(), sg.FileBrowse(key="exlo")],
            [sg.Text('Enter the type of Automation you want')],
            [sg.Listbox(values=['Banner Copy Download', 'Data Scrubbing', 'Native file Download', 'CO Attachments download'], select_mode='extended', key='fac', size=(30, 6))],
            [sg.Button('Ok'), sg.Button('Cancel')] ]
window = sg.Window('Automatic Scripts', layout)
event, values = window.read()
window.close()
wb = openpyxl.Workbook()    
ws = wb.active
Filename = pd.read_excel(values["exlo"])
l=Filename['File_Name'].tolist()
listToStr = ' '.join(map(str, values["fac"]))
wb = openpyxl.Workbook()
ws = wb.active
inst = webdriver.Chrome()
inst.get('https://agileprod.jnj.com/Agile/default/login-cms.jsp')
window_before = inst.window_handles[0]
usr=inst.find_element('id','j_username')
usr.send_keys(values[0])
pas=inst.find_element('id','j_password')
pas.send_keys(values[1])
inst.find_element('id','login').click()
for i in l:
    err=[]
    window_after = inst.window_handles[1]
    inst.switch_to.window(window_after)
    inst.implicitly_wait(2)
    WebDriverWait(inst, 20).until(EC.element_to_be_clickable((By.ID, 'QUICKSEARCH_STRING')))
    inst.find_element('id','QUICKSEARCH_STRING').clear()
    search=inst.find_element('id','QUICKSEARCH_STRING')
    search.send_keys(i)
    inst.find_element(By.ID, "selector_elm").click()
    inst.find_element(By.LINK_TEXT, "Items").click()
    inst.find_element(By.CSS_SELECTOR, ".quick_search").click()
    try:
        inst.find_element(By.ID, "col_1001")
    except:
        try:
            inst.find_element(By.LINK_TEXT,i).click()
        except:
            err.append(i)
            err.append("Label not found")
            ws.append(err)
            wb.save('Output.xlsx')
            pass
            break
    if listToStr == "Banner Copy Download":
        bcd(i,values["dl"])
    elif listToStr == "Data Scrubbing":
        a=scrap()
        b=[]
        b=a.scr()
        ws.append(b)
        wb.save('Ouput.xlsx')
    elif listToStr == "Native file Download": 
        nfd()
inst.implicitly_wait(5)
layout1 = [  [sg.Text('Program Complete')],
            [sg.Button('Ok')] ]
window1 = sg.Window('Automatic Scripts', layout1)
newevent,newvalues = window1.read()
window1.close()
if newevent == "Ok":
    inst.quit()





